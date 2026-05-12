"""Pure-function tests for header anchor resolution. v1.16.0."""

from __future__ import annotations

from openpyxl import Workbook

from aa_auto_sdr.output._template_anchors import (
    ANCHORS,
    ResolvedSheet,
    SheetAnchor,
    resolve_sheet,
)


def _build_workbook(
    *,
    sheet_name: str,
    section_title_b4: str | None,
    header_row: int | None,
    headers: list[str] | None = None,
) -> Workbook:
    """Synthetic workbook helper. `header_row=None` means no ID marker anywhere."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    if section_title_b4 is not None:
        ws["B4"] = section_title_b4
    if header_row is not None and headers is not None:
        for col_idx, hdr in enumerate(headers, start=2):  # column B = 2
            ws.cell(row=header_row, column=col_idx, value=hdr)
    return wb


def test_missing_sheet_returns_none() -> None:
    wb = Workbook()
    wb.active.title = "other"  # type: ignore[union-attr]
    assert resolve_sheet(wb, SheetAnchor("eVars", "eVars")) is None


def test_wrong_b4_section_title_returns_none() -> None:
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="Props",  # wrong
        header_row=6,
        headers=["ID", "Analytics Variable", "Variable Name"],
    )
    assert resolve_sheet(wb, SheetAnchor("eVars", "eVars")) is None


def test_b4_section_title_case_insensitive_match() -> None:
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="EVARS",
        header_row=6,
        headers=["ID", "Analytics Variable", "Variable Name"],
    )
    rs = resolve_sheet(wb, SheetAnchor("eVars", "eVars"))
    assert rs is not None
    assert rs.header_row == 6


def test_b4_section_title_with_padding_match() -> None:
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="  eVars  ",
        header_row=6,
        headers=["ID", "Analytics Variable"],
    )
    rs = resolve_sheet(wb, SheetAnchor("eVars", "eVars"))
    assert rs is not None


def test_no_id_marker_in_rows_5_to_10_returns_none() -> None:
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="eVars",
        header_row=6,
        headers=["NotID", "Analytics Variable"],  # B6 != "ID"
    )
    assert resolve_sheet(wb, SheetAnchor("eVars", "eVars")) is None


def test_header_at_row_5_resolved() -> None:
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="eVars",
        header_row=5,
        headers=["ID", "Analytics Variable", "Variable Name"],
    )
    rs = resolve_sheet(wb, SheetAnchor("eVars", "eVars"))
    assert rs == ResolvedSheet(
        sheet_name="eVars",
        header_row=5,
        columns={"ID": 2, "Analytics Variable": 3, "Variable Name": 4},
        first_data_row=6,
    )


def test_header_at_row_10_resolved() -> None:
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="eVars",
        header_row=10,
        headers=["ID", "Analytics Variable"],
    )
    rs = resolve_sheet(wb, SheetAnchor("eVars", "eVars"))
    assert rs is not None
    assert rs.header_row == 10
    assert rs.first_data_row == 11


def test_header_at_row_11_not_searched() -> None:
    """Per §3.4, scan rows 5–10 only."""
    wb = _build_workbook(
        sheet_name="eVars",
        section_title_b4="eVars",
        header_row=11,
        headers=["ID", "Analytics Variable"],
    )
    assert resolve_sheet(wb, SheetAnchor("eVars", "eVars")) is None


def test_columns_skip_blank_cells_in_header_row() -> None:
    """A blank cell mid-header shouldn't poison the column map."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "eVars"
    ws["B4"] = "eVars"
    ws["B6"] = "ID"
    ws["C6"] = "Analytics Variable"
    ws["D6"] = None  # blank header cell
    ws["E6"] = "Variable Name"
    rs = resolve_sheet(wb, SheetAnchor("eVars", "eVars"))
    assert rs is not None
    assert rs.columns == {"ID": 2, "Analytics Variable": 3, "Variable Name": 5}


def test_anchors_dict_has_four_data_sheet_entries() -> None:
    """Per §3.5 — Glossary handled separately, four data sheets."""
    assert set(ANCHORS.keys()) == {
        "evars",
        "props",
        "events",
        "metrics_segments",
    }
    assert ANCHORS["events"].sheet_name == "custom events (metrics)"
    assert ANCHORS["events"].section_title == "Custom Events (Metrics)"
    assert ANCHORS["metrics_segments"].sheet_name == "metrics-segments"
    assert ANCHORS["metrics_segments"].section_title == "Metrics - Segments"
