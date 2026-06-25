"""Targeted coverage for ExcelTemplateWriter skip/append/soft-cap branches.

Exercises the engine methods directly against synthetic openpyxl workbooks
(no Adobe template needed): missing/unanchored sheets, missing id columns,
None-id rows, the +50 soft cap with row drops, the Glossary-absent guard,
and the non-empty rule in `_write_row`."""

from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import Workbook

from aa_auto_sdr.api import models
from aa_auto_sdr.output._template_anchors import ANCHORS
from aa_auto_sdr.output.writers.excel_template import ExcelTemplateWriter, _write_row
from aa_auto_sdr.sdr.document import SdrDocument


def _add_resolvable_sheet(wb: Workbook, *, sheet_name: str, section_title: str, headers: list[str]) -> None:
    """Create a sheet that `resolve_sheet` will anchor: B4 section title plus
    an `ID`-led header row at row 6."""
    ws = wb.create_sheet(sheet_name)
    ws["B4"] = section_title
    for col_idx, hdr in enumerate(headers, start=2):  # column B = 2
        ws.cell(row=6, column=col_idx, value=hdr)


def _dim(id_: str, name: str, description: str | None = None) -> models.Dimension:
    return models.Dimension(
        id=id_,
        name=name,
        type="string",
        category=None,
        parent="",
        pathable=False,
        description=description,
    )


def _seg(i: int) -> models.Segment:
    return models.Segment(
        id=f"seg_{i}",
        name=f"Segment {i}",
        description=f"desc {i}",
        rsid="rs",
        owner_id=None,
        definition={},
    )


def _make_doc(*, segments: list[models.Segment] | None = None) -> SdrDocument:
    return SdrDocument(
        report_suite=models.ReportSuite(
            rsid="rs",
            name="RS Name",
            timezone=None,
            currency=None,
            parent_rsid=None,
        ),
        dimensions=[],
        metrics=[],
        segments=segments or [],
        calculated_metrics=[],
        virtual_report_suites=[],
        classifications=[],
        captured_at=datetime(2026, 5, 12, tzinfo=UTC),
        tool_version="1.16.0",
    )


# --- _fill_by_id_sheet ----------------------------------------------------


def test_fill_by_id_sheet_skips_when_sheet_missing_or_unanchored() -> None:
    wb = Workbook()  # only the default "Sheet"; no "eVars"
    writer = ExcelTemplateWriter()
    # resolve_sheet returns None → warn + early return, no exception.
    writer._fill_by_id_sheet(
        wb,
        anchor=ANCHORS["evars"],
        id_column_header="Analytics Variable",
        components=[],
        get_id=lambda d: d.id,
        column_map={},
    )
    assert "eVars" not in wb.sheetnames


def test_fill_by_id_sheet_skips_when_id_column_absent() -> None:
    wb = Workbook()
    _add_resolvable_sheet(
        wb,
        sheet_name="eVars",
        section_title="eVars",
        headers=["ID", "Analytics Variable", "Variable Name", "Variable Description"],
    )
    writer = ExcelTemplateWriter()
    # Resolves, but the requested id column header is not in the sheet → return.
    writer._fill_by_id_sheet(
        wb,
        anchor=ANCHORS["evars"],
        id_column_header="Nonexistent Column",
        components=[_dim("evar1", "Customer ID")],
        get_id=lambda d: d.id,
        column_map={"Variable Name": lambda d: d.name},
    )
    # Nothing written into the data region.
    ws = wb["eVars"]
    assert ws.cell(row=7, column=4).value is None


def test_fill_by_id_sheet_skips_rows_with_none_id() -> None:
    wb = Workbook()
    _add_resolvable_sheet(
        wb,
        sheet_name="eVars",
        section_title="eVars",
        headers=["ID", "Analytics Variable", "Variable Name", "Variable Description"],
    )
    ws = wb["eVars"]
    # Put a value in a non-id column at row 8 so max_row advances while the
    # id column (column 3) stays None — exercises the None-id skip.
    ws.cell(row=8, column=5, value="orphan note")
    writer = ExcelTemplateWriter()
    writer._fill_by_id_sheet(
        wb,
        anchor=ANCHORS["evars"],
        id_column_header="Analytics Variable",
        components=[_dim("evar1", "Customer ID")],
        get_id=lambda d: d.id,
        column_map={"Variable Name": lambda d: d.name},
    )
    # evar1 had no matching row → appended below the None-id rows.
    seen = [ws.cell(row=r, column=3).value for r in range(7, ws.max_row + 1)]
    assert "evar1" in seen


def test_fill_by_id_sheet_soft_cap_breaks_and_drops_overflow() -> None:
    wb = Workbook()
    _add_resolvable_sheet(
        wb,
        sheet_name="eVars",
        section_title="eVars",
        headers=["ID", "Analytics Variable", "Variable Name", "Variable Description"],
    )
    writer = ExcelTemplateWriter()
    # 60 components, no template rows to match → all go to the append path.
    # Soft cap is max_row(6) + 50 = 56, so only 50 append and 10 are dropped.
    components = [f"x{i}" for i in range(60)]
    writer._fill_by_id_sheet(
        wb,
        anchor=ANCHORS["evars"],
        id_column_header="Analytics Variable",
        components=components,
        get_id=lambda c: c,
        column_map={},
    )
    ws = wb["eVars"]
    appended_ids = {ws.cell(row=r, column=3).value for r in range(7, ws.max_row + 1)}
    appended_ids.discard(None)
    assert len(appended_ids) == 50  # soft cap stopped at +50 rows


# --- _fill_metrics_segments ----------------------------------------------


def test_fill_metrics_segments_skips_when_sheet_missing() -> None:
    wb = Workbook()  # no "metrics-segments" sheet
    writer = ExcelTemplateWriter()
    # resolve_sheet returns None → warn + early return.
    writer._fill_metrics_segments(wb, _make_doc(segments=[_seg(0)]))
    assert "metrics-segments" not in wb.sheetnames


def test_fill_metrics_segments_soft_cap_breaks_and_drops_overflow() -> None:
    wb = Workbook()
    _add_resolvable_sheet(
        wb,
        sheet_name="metrics-segments",
        section_title="Metrics - Segments",
        headers=["ID", "Type", "Name", "Description", "Format"],
    )
    writer = ExcelTemplateWriter()
    # 60 entries, append-only sheet, soft cap = max_row(6) + 50 = 56 → 50 land.
    writer._fill_metrics_segments(wb, _make_doc(segments=[_seg(i) for i in range(60)]))
    ws = wb["metrics-segments"]
    names = {ws.cell(row=r, column=4).value for r in range(7, ws.max_row + 1)}
    names.discard(None)
    assert len(names) == 50


# --- _fill_glossary_org ---------------------------------------------------


def test_fill_glossary_org_skips_when_glossary_absent() -> None:
    wb = Workbook()  # default "Sheet"; no "Glossary"
    writer = ExcelTemplateWriter()
    writer._fill_glossary_org(wb, _make_doc())  # early return, no exception
    assert "Glossary" not in wb.sheetnames


def test_fill_glossary_org_writes_org_override_to_c2() -> None:
    wb = Workbook()
    wb.create_sheet("Glossary")
    writer = ExcelTemplateWriter()
    writer.organization = "Acme Corp"
    writer._fill_glossary_org(wb, _make_doc())
    assert wb["Glossary"]["C2"].value == "Acme Corp"


def test_fill_glossary_org_defaults_to_report_suite_name() -> None:
    wb = Workbook()
    wb.create_sheet("Glossary")
    writer = ExcelTemplateWriter()
    writer._fill_glossary_org(wb, _make_doc())
    assert wb["Glossary"]["C2"].value == "RS Name"


# --- _write_row -----------------------------------------------------------


def test_write_row_skips_header_not_in_columns() -> None:
    wb = Workbook()
    ws = wb.active
    _write_row(ws, row=1, columns={"Name": 2}, column_map={"Missing": lambda _c: "value"}, comp=object())
    assert ws.cell(row=1, column=2).value is None  # nothing written for absent column


def test_write_row_skips_whitespace_only_value() -> None:
    wb = Workbook()
    ws = wb.active
    _write_row(ws, row=1, columns={"Name": 2}, column_map={"Name": lambda _c: "   "}, comp=object())
    assert ws.cell(row=1, column=2).value is None  # blank value never overwrites
