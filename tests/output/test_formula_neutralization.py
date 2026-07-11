"""Formula-injection neutralization in tabular writers.

Component names are authored by any org user in Adobe Analytics. A name like
`=HYPERLINK(...)` must never become a live formula when the SDR is opened in
Excel — neither via the CSV writer (Excel evaluates on import) nor via the
xlsxwriter/openpyxl paths (both convert leading-`=` strings to formula cells
by default).
"""

from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from aa_auto_sdr.api import models
from aa_auto_sdr.output.writers.csv import CsvWriter
from aa_auto_sdr.output.writers.excel import ExcelWriter
from aa_auto_sdr.output.writers.excel_template import _write_row
from aa_auto_sdr.sdr.document import SdrDocument

INJECTED = '=HYPERLINK("http://evil.example","click")'


def _dim(name: str) -> models.Dimension:
    return models.Dimension(
        id="evar1",
        name=name,
        type="dimension",
        category=None,
        parent="",
        pathable=False,
        description=None,
    )


def _doc(dim_name: str) -> SdrDocument:
    rs = models.ReportSuite(rsid="rs1", name="RS1", timezone=None, currency=None, parent_rsid=None)
    return SdrDocument(
        report_suite=rs,
        dimensions=[_dim(dim_name)],
        metrics=[],
        segments=[],
        calculated_metrics=[],
        virtual_report_suites=[],
        classifications=[],
        captured_at=datetime(2026, 7, 11, tzinfo=UTC),
        tool_version="1.21.6",
    )


class TestCsvWriterNeutralizesFormulas:
    def test_leading_equals_is_neutralized(self, tmp_path: Path) -> None:
        CsvWriter().write(_doc(INJECTED), tmp_path / "out.csv")
        with (tmp_path / "out.dimensions.csv").open(encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        header, row = rows[0], rows[1]
        name_cell = row[header.index("name")]
        assert not name_cell.startswith("=")
        assert INJECTED in name_cell  # payload preserved, just defused

    def test_plain_names_unchanged(self, tmp_path: Path) -> None:
        CsvWriter().write(_doc("Page Name"), tmp_path / "out.csv")
        with (tmp_path / "out.dimensions.csv").open(encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        header, row = rows[0], rows[1]
        assert row[header.index("name")] == "Page Name"


class TestExcelWriterNeutralizesFormulas:
    def test_component_cell_is_text_not_formula(self, tmp_path: Path) -> None:
        target = tmp_path / "sdr.xlsx"
        ExcelWriter().write(_doc(INJECTED), target)
        wb = load_workbook(target)
        ws = wb["Dimensions"]
        headers = [c.value for c in ws[1]]
        name_col = headers.index("name") + 1
        cell = ws.cell(row=2, column=name_col)
        assert cell.data_type != "f", "component name was written as a live formula"
        assert cell.value == INJECTED

    def test_summary_name_cell_is_text_not_formula(self, tmp_path: Path) -> None:
        doc = _doc("x")
        doc = SdrDocument(
            report_suite=models.ReportSuite(rsid="rs1", name=INJECTED, timezone=None, currency=None, parent_rsid=None),
            dimensions=doc.dimensions,
            metrics=[],
            segments=[],
            calculated_metrics=[],
            virtual_report_suites=[],
            classifications=[],
            captured_at=doc.captured_at,
            tool_version=doc.tool_version,
        )
        target = tmp_path / "sdr.xlsx"
        ExcelWriter().write(doc, target)
        ws = load_workbook(target)["Summary"]
        formula_cells = [c.coordinate for row in ws.iter_rows() for c in row if c.data_type == "f"]
        assert formula_cells == []


class TestTemplateWriterNeutralizesFormulas:
    def test_write_row_keeps_leading_equals_as_text(self) -> None:
        wb = Workbook()
        ws = wb.active
        _write_row(
            ws,
            row=2,
            columns={"Name": 1},
            column_map={"Name": lambda c: c.name},
            comp=_dim(INJECTED),
        )
        cell = ws.cell(row=2, column=1)
        assert cell.data_type != "f", "template fill wrote a live formula"
        assert str(cell.value).endswith(INJECTED)
