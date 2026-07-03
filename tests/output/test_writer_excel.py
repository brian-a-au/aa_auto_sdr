"""Excel writer: one sheet per component type, frozen header row, autofilter."""

import json
from datetime import UTC, datetime
from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import load_workbook

from aa_auto_sdr.api.client import AaClient
from aa_auto_sdr.output.writers.excel import ExcelWriter
from aa_auto_sdr.sdr.builder import build_sdr

FIXTURE = Path(__file__).parent.parent / "fixtures" / "sample_rs.json"


def _df(records: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(records)


@pytest.fixture
def doc():
    raw = json.loads(FIXTURE.read_text())
    handle = MagicMock()
    handle.getReportSuites.return_value = _df([raw["report_suite"]])
    handle.getDimensions.return_value = _df(raw["dimensions"])
    handle.getMetrics.return_value = _df(raw["metrics"])
    handle.getSegments.return_value = _df(raw["segments"])
    handle.getCalculatedMetrics.return_value = _df(raw["calculated_metrics"])
    handle.getVirtualReportSuites.return_value = _df(raw["virtual_report_suites"])
    handle.getClassificationDatasets.return_value = _df(raw["classification_datasets"])
    client = AaClient(handle=handle, company_id="testco")
    return build_sdr(client, "demo.prod", captured_at=datetime(2026, 4, 25, tzinfo=UTC), tool_version="0.1.0")


def test_excel_extension() -> None:
    assert ExcelWriter().extension == ".xlsx"


def test_excel_writer_creates_file(doc, tmp_path: Path) -> None:
    target = tmp_path / "sdr.xlsx"
    actual = ExcelWriter().write(doc, target)
    assert actual == [target]
    assert target.exists()


def test_excel_has_summary_and_one_sheet_per_component(doc, tmp_path: Path) -> None:
    target = tmp_path / "sdr.xlsx"
    ExcelWriter().write(doc, target)
    wb = load_workbook(target, read_only=True)
    expected = {
        "Summary",
        "Dimensions",
        "Metrics",
        "Segments",
        "Calculated Metrics",
        "Virtual Report Suites",
        "Classifications",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_excel_freezes_header_row(doc, tmp_path: Path) -> None:
    target = tmp_path / "sdr.xlsx"
    ExcelWriter().write(doc, target)
    wb = load_workbook(target, read_only=False)
    assert wb["Dimensions"].freeze_panes == "A2"


def test_excel_dimension_rows_match_doc(doc, tmp_path: Path) -> None:
    target = tmp_path / "sdr.xlsx"
    ExcelWriter().write(doc, target)
    wb = load_workbook(target, read_only=True)
    sheet = wb["Dimensions"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "id"
    assert len(rows) == 1 + len(doc.dimensions)


def test_excel_column_widths_fit_content_capped_at_60(tmp_path: Path) -> None:
    """Column width = max(header, longest cell) + 2, capped at 60.

    Characterization for the width pass — locks rendered output across the
    single-stringify refactor (widths are computed from the already-
    stringified cells, not a second astype(str) conversion)."""
    from dataclasses import replace as dc_replace

    raw = json.loads(FIXTURE.read_text())
    handle = MagicMock()
    handle.getReportSuites.return_value = _df([raw["report_suite"]])
    long_desc = "x" * 100
    handle.getDimensions.return_value = _df(
        [
            {"id": "evar1", "name": "A", "type": "string", "description": long_desc, "tags": []},
            {"id": "evar2", "name": "B", "type": "string", "description": "short", "tags": []},
        ],
    )
    handle.getMetrics.return_value = _df(raw["metrics"])
    handle.getSegments.return_value = _df(raw["segments"])
    handle.getCalculatedMetrics.return_value = _df(raw["calculated_metrics"])
    handle.getVirtualReportSuites.return_value = _df(raw["virtual_report_suites"])
    handle.getClassificationDatasets.return_value = _df(raw["classification_datasets"])
    client = AaClient(handle=handle, company_id="testco")
    doc = build_sdr(client, "demo.prod", captured_at=datetime(2026, 4, 25, tzinfo=UTC), tool_version="0.1.0")
    doc = dc_replace(doc)

    target = tmp_path / "sdr.xlsx"
    ExcelWriter().write(doc, target)
    wb = load_workbook(target)
    ws = wb["Dimensions"]
    headers = [c.value for c in ws[1]]
    id_col = chr(ord("A") + headers.index("id"))
    desc_col = chr(ord("A") + headers.index("description"))
    # xlsxwriter stores character widths with sub-unit font padding, so
    # openpyxl reads e.g. 7 as ~7.71 — assert within the one-unit band.
    # id: max(len("id")=2, len("evar1")=5) + 2 = 7
    assert 7 <= ws.column_dimensions[id_col].width < 8
    # description: 100-char cell exceeds the cap -> 60
    assert 60 <= ws.column_dimensions[desc_col].width < 61
