"""ExcelTemplateWriter integration tests against a synthetic template. v1.16.0."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from aa_auto_sdr.api import models
from aa_auto_sdr.output import registry
from aa_auto_sdr.output.writers.excel_template import ExcelTemplateWriter
from aa_auto_sdr.sdr.document import SdrDocument


def _doc(
    *,
    rsid: str = "test_rs",
    dimensions: list[models.Dimension] | None = None,
    metrics: list[models.Metric] | None = None,
    segments: list[models.Segment] | None = None,
    calculated_metrics: list[models.CalculatedMetric] | None = None,
) -> SdrDocument:
    return SdrDocument(
        report_suite=models.ReportSuite(
            rsid=rsid,
            name="Test Report Suite",
            timezone="UTC",
            currency=None,
            parent_rsid=None,
        ),
        dimensions=dimensions or [],
        metrics=metrics or [],
        segments=segments or [],
        calculated_metrics=calculated_metrics or [],
        virtual_report_suites=[],
        classifications=[],
        captured_at=datetime(2026, 5, 12, 0, 0, 0, tzinfo=UTC),
        tool_version="1.16.0",
    )


def test_writer_registers_under_excel_template_key() -> None:
    registry.bootstrap()
    assert registry.get_writer("excel-template").extension == ".xlsx"


def test_write_raises_without_template_path(tmp_path: Path) -> None:
    writer = ExcelTemplateWriter()
    with pytest.raises(RuntimeError, match="template_path"):
        writer.write(_doc(), tmp_path / "out.xlsx")


def test_write_round_trip_preserves_untouched_cells(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    paths = writer.write(_doc(), out)
    assert paths == [out]
    assert out.exists()
    wb = load_workbook(out)
    assert wb["eVars"]["C1"].value == "Adobe Analytics"
    assert wb["eVars"]["C2"].value == "=Glossary!C2"
    assert wb["eVars"]["B4"].value == "eVars"


def test_glossary_c2_defaults_to_report_suite_name(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(_doc(rsid="my_rs"), out)
    wb = load_workbook(out)
    assert wb["Glossary"]["C2"].value == "Test Report Suite"


def test_glossary_c2_uses_organization_override(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    writer.organization = "Acme Corp"
    out = tmp_path / "out.xlsx"
    writer.write(_doc(), out)
    wb = load_workbook(out)
    assert wb["Glossary"]["C2"].value == "Acme Corp"


def test_cross_sheet_formula_still_present_after_glossary_write(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    """The cross-sheet formula on every non-Glossary C2 must survive."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(_doc(), out)
    wb = load_workbook(out)
    for sheet in ("eVars", "props", "custom events (metrics)", "metrics-segments"):
        assert wb[sheet]["C2"].value == "=Glossary!C2", sheet


def _dim(id_: str, name: str, description: str | None = None) -> models.Dimension:
    return models.Dimension(
        id=id_,
        name=name,
        type="string",
        category=None,
        parent="",
        pathable=False,
        description=description,
    )


def test_dimension_match_by_id_overwrites_preseeded_evars(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    dims = [
        _dim("evar1", "Customer ID", "Logged-in customer hash"),
        _dim("evar2", "Cart ID", "Active cart token"),
    ]
    writer.write(_doc(dimensions=dims), out)
    wb = load_workbook(out)
    ws = wb["eVars"]
    # eVar1 at row 7 was pre-seeded; should be overwritten
    assert ws.cell(row=7, column=3).value == "evar1"  # Analytics Variable col
    assert ws.cell(row=7, column=4).value == "Customer ID"
    assert ws.cell(row=7, column=5).value == "Logged-in customer hash"
    # eVar2 at row 8
    assert ws.cell(row=8, column=4).value == "Cart ID"


def test_dimension_skeleton_rows_with_no_api_match_untouched(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    """eVar37 is a template-only skeleton row; API has no eVar37 → leave it."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(_doc(dimensions=[_dim("evar1", "Customer ID")]), out)
    wb = load_workbook(out)
    ws = wb["eVars"]
    # Row 9 was eVar37 skeleton; untouched
    assert ws.cell(row=9, column=3).value == "evar37"
    assert ws.cell(row=9, column=4).value == "Skeleton eVar 37"


def test_dimension_match_is_case_insensitive(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    """Template has 'evar1'; API returns 'eVar1' — must still match."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(_doc(dimensions=[_dim("eVar1", "Customer ID")]), out)
    wb = load_workbook(out)
    ws = wb["eVars"]
    assert ws.cell(row=7, column=4).value == "Customer ID"


def test_dimension_blank_description_does_not_overwrite_existing(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    """Rule §3.7: don't blank-overwrite when our value is None/empty."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(_doc(dimensions=[_dim("evar1", "Customer ID", description=None)]), out)
    wb = load_workbook(out)
    ws = wb["eVars"]
    # Description column should keep the pre-seeded value, not become blank.
    assert ws.cell(row=7, column=5).value == "Pre-seeded desc 1"


def test_props_match_by_id(synthetic_template_path: Path, tmp_path: Path) -> None:
    """pageName is reserved/pre-seeded; API match should overwrite."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(
        _doc(
            dimensions=[
                _dim("prop1", "Internal Section"),
                _dim("pageName", "Page Name", "Canonical URL path"),
            ],
        ),
        out,
    )
    wb = load_workbook(out)
    ws = wb["props"]
    # prop1 at row 7
    assert ws.cell(row=7, column=4).value == "Internal Section"
    # pageName at row 8 (reserved, overwritten)
    assert ws.cell(row=8, column=4).value == "Page Name"
    assert ws.cell(row=8, column=5).value == "Canonical URL path"


def test_evars_sheet_filtering_includes_campaign_excludes_prop(
    synthetic_template_with_campaign_skeleton: Path,
    tmp_path: Path,
) -> None:
    """eVars sheet gets evar*-prefixed dims + 'campaign'; props get prop*+pageName/linkName."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_with_campaign_skeleton
    out = tmp_path / "out.xlsx"
    writer.write(
        _doc(
            dimensions=[
                _dim("campaign", "Tracking Code"),
                _dim("prop1", "Internal Section"),  # should not appear in eVars sheet
            ],
        ),
        out,
    )
    wb = load_workbook(out)
    # campaign overwrites the pre-seeded row at eVars row 10
    assert wb["eVars"].cell(row=10, column=3).value == "campaign"
    assert wb["eVars"].cell(row=10, column=4).value == "Tracking Code"
    # prop1 lands in props sheet, not eVars
    assert wb["props"].cell(row=7, column=4).value == "Internal Section"


def test_dimension_filter_case_insensitive_on_reserved_ids(
    synthetic_template_with_campaign_skeleton: Path,
    tmp_path: Path,
) -> None:
    """Reserved ids (campaign, pageName, linkName) match case-insensitively."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_with_campaign_skeleton
    out = tmp_path / "out.xlsx"
    writer.write(
        _doc(
            dimensions=[
                _dim("CAMPAIGN", "Mixed Case Campaign"),
                _dim("PAGENAME", "Mixed Case Page Name"),
            ],
        ),
        out,
    )
    wb = load_workbook(out)
    # CAMPAIGN matches the lowercase 'campaign' skeleton row at eVars row 10
    assert wb["eVars"].cell(row=10, column=3).value in ("campaign", "CAMPAIGN")
    assert wb["eVars"].cell(row=10, column=4).value == "Mixed Case Campaign"
    # PAGENAME matches the lowercase 'pageName' skeleton row at props row 8
    assert wb["props"].cell(row=8, column=4).value == "Mixed Case Page Name"


def _metric(id_: str, name: str, description: str | None = None) -> models.Metric:
    return models.Metric(
        id=id_,
        name=name,
        type="counter",
        category=None,
        precision=0,
        segmentable=True,
        description=description,
    )


def test_metrics_fill_custom_events_sheet(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(
        _doc(
            metrics=[
                _metric("event1", "Add to Cart", "Click on the add-to-cart button"),
                _metric("event5", "Checkout Start", None),
            ],
        ),
        out,
    )
    wb = load_workbook(out)
    ws = wb["custom events (metrics)"]
    # event1 at row 7 was pre-seeded; should overwrite
    assert ws.cell(row=7, column=3).value == "event1"
    assert ws.cell(row=7, column=4).value == "Add to Cart"
    assert ws.cell(row=7, column=5).value == "Click on the add-to-cart button"


def test_metrics_exclude_built_in_visitors(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    """Built-in metrics like 'visitors' / 'pageViews' must not be written
    to the events sheet."""
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(
        _doc(
            metrics=[
                _metric("visitors", "Visitors"),
                _metric("event1", "Add to Cart"),
            ],
        ),
        out,
    )
    wb = load_workbook(out)
    ws = wb["custom events (metrics)"]
    # event1 written; 'visitors' must NOT appear anywhere on this sheet.
    seen = {ws.cell(row=r, column=3).value for r in range(7, ws.max_row + 1)}
    assert "event1" in seen
    assert "visitors" not in seen


def _calc(id_: str, name: str, precision: int = 0, type_: str = "decimal") -> models.CalculatedMetric:
    return models.CalculatedMetric(
        id=id_,
        name=name,
        description=f"Calc metric {name}",
        rsid="test_rs",
        owner_id=None,
        polarity="positive",
        precision=precision,
        type=type_,
        definition={},
    )


def _seg(id_: str, name: str) -> models.Segment:
    return models.Segment(
        id=id_,
        name=name,
        description=f"Segment {name}",
        rsid="test_rs",
        owner_id=None,
        definition={},
    )


def test_metrics_segments_appends_calc_metrics_with_type_column(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    writer.write(
        _doc(
            calculated_metrics=[_calc("cm_1", "Conversion Rate", type_="percent")],
            segments=[_seg("seg_1", "US Visitors")],
        ),
        out,
    )
    wb = load_workbook(out)
    ws = wb["metrics-segments"]
    # No pre-seeded rows; both entries appear in the append range.
    rows: list[tuple[str | None, str | None, str | None, str | None]] = [
        (
            ws.cell(row=r, column=3).value,  # Type
            ws.cell(row=r, column=4).value,  # Name
            ws.cell(row=r, column=5).value,  # Description
            ws.cell(row=r, column=6).value,  # Format
        )
        for r in range(7, ws.max_row + 1)
    ]
    names = [n for (_t, n, _d, _f) in rows if n]
    assert "Conversion Rate" in names
    assert "US Visitors" in names
    # Type column populated:
    types = {t for (t, n, _d, _f) in rows if n}
    assert "Calculated Metric" in types
    assert "Segment" in types
    # Format column only populated for the calc metric (percent), blank for segment.
    cm_row = next(r for r in rows if r[1] == "Conversion Rate")
    seg_row = next(r for r in rows if r[1] == "US Visitors")
    assert cm_row[3] == "Percent"
    assert seg_row[3] in (None, "")


def test_dimension_append_path_extends_past_max_row(
    synthetic_template_path: Path,
    tmp_path: Path,
    caplog: pytest.LogCaptureFixture,
) -> None:
    """An API-only id with no template skeleton row should append a new row."""
    import logging as _logging

    writer = ExcelTemplateWriter()
    writer.template_path = synthetic_template_path
    out = tmp_path / "out.xlsx"
    caplog.set_level(_logging.WARNING, logger="aa_auto_sdr.output.writers.excel_template")
    writer.write(
        _doc(
            dimensions=[
                _dim("evar1", "Customer ID"),  # match path
                _dim("evar99", "API-only eVar"),  # append path
            ],
        ),
        out,
    )
    wb = load_workbook(out)
    ws = wb["eVars"]
    # Walk rows looking for evar99 (must exist past the original max_row).
    found_evar99 = False
    for r in range(7, ws.max_row + 1):
        if ws.cell(row=r, column=3).value == "evar99":
            assert ws.cell(row=r, column=4).value == "API-only eVar"
            found_evar99 = True
    assert found_evar99
    # Overflow warning emitted (any rows_appended > 0 emits it per §3.7).
    overflow_records = [r for r in caplog.records if r.message.startswith("template_overflow")]
    assert overflow_records, "expected template_overflow WARNING"
    rec = overflow_records[0]
    assert getattr(rec, "sheet", None) == "eVars"
    assert getattr(rec, "overflow_rows", 0) >= 1


def test_pipeline_writer_instance_is_configured_via_attribute(
    synthetic_template_path: Path,
    tmp_path: Path,
) -> None:
    """Pipeline boundary: setting `template_path` + `organization` on the
    registered instance (no protocol change) is the supported config path."""
    registry.bootstrap()
    writer = registry.get_writer("excel-template")
    writer.template_path = synthetic_template_path
    writer.organization = "Pipeline Test Org"
    try:
        paths = writer.write(_doc(), tmp_path / "piped.xlsx")
        wb = load_workbook(paths[0])
        assert wb["Glossary"]["C2"].value == "Pipeline Test Org"
    finally:
        writer.template_path = None
        writer.organization = None
