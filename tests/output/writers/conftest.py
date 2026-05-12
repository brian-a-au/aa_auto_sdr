"""Synthetic Adobe BRD/SDR template fixture for ExcelTemplateWriter tests.

Builds a minimal `.xlsx` programmatically that mimics the real template's
layout (Glossary + 4 data sheets, B4 section titles, B6 headers, a few
skeleton rows, a `=Glossary!C2` formula on each non-Glossary sheet). This
decouples the test suite from any specific Adobe template version and avoids
redistribution concerns. v1.16.0."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook


def _add_data_sheet(
    wb: Workbook,
    *,
    sheet_name: str,
    section_title: str,
    headers: list[str],
    skeleton_rows: list[list[str]],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws["C1"] = "Adobe Analytics"  # brand banner
    ws["C2"] = "=Glossary!C2"  # cross-sheet formula
    ws["B4"] = section_title  # section title
    for col_idx, hdr in enumerate(headers, start=2):  # column B = 2
        ws.cell(row=6, column=col_idx, value=hdr)
    for row_offset, row in enumerate(skeleton_rows, start=7):
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_offset, column=col_idx, value=value)


def _build_template(
    path: Path,
    *,
    evars_skeleton: list[list[str]] | None = None,
) -> Path:
    """Build the synthetic Adobe-shaped template at `path` and return the path.

    `evars_skeleton` lets callers vary the eVars sheet's pre-seeded rows
    without mutating an already-built fixture file. Defaults to the standard
    three rows (evar1, evar2, evar37)."""
    wb = Workbook()
    glossary = wb.active
    assert glossary is not None
    glossary.title = "Glossary"
    glossary["C1"] = "Adobe Analytics"
    glossary["C2"] = "Customer Org Placeholder"  # the source-of-truth cell

    _add_data_sheet(
        wb,
        sheet_name="eVars",
        section_title="eVars",
        headers=["ID", "Analytics Variable", "Variable Name", "Variable Description"],
        skeleton_rows=evars_skeleton
        if evars_skeleton is not None
        else [
            ["1", "evar1", "Pre-seeded eVar 1", "Pre-seeded desc 1"],
            ["2", "evar2", "Pre-seeded eVar 2", "Pre-seeded desc 2"],
            ["3", "evar37", "Skeleton eVar 37", "Skeleton desc 37"],
        ],
    )
    _add_data_sheet(
        wb,
        sheet_name="props",
        section_title="Props",
        headers=["ID", "Analytics Variable", "Variable Name", "Variable Description"],
        skeleton_rows=[
            ["1", "prop1", "Pre-seeded prop 1", "Pre-seeded prop desc 1"],
            ["2", "pageName", "Reserved pageName", "Reserved page-name desc"],
        ],
    )
    _add_data_sheet(
        wb,
        sheet_name="custom events (metrics)",
        section_title="Custom Events (Metrics)",
        headers=["ID", "Event", "Event Name", "Event Description"],
        skeleton_rows=[
            ["1", "event1", "Pre-seeded event 1", "Pre-seeded event desc 1"],
            ["2", "event20", "Skeleton event 20", "Skeleton event desc 20"],
        ],
    )
    _add_data_sheet(
        wb,
        sheet_name="metrics-segments",
        section_title="Metrics - Segments",
        headers=["ID", "Type", "Name", "Description", "Format"],
        skeleton_rows=[],  # empty start
    )
    wb.save(path)
    return path


@pytest.fixture
def synthetic_template_path(tmp_path: Path) -> Path:
    return _build_template(tmp_path / "synthetic_template.xlsx")


@pytest.fixture
def synthetic_template_with_campaign_skeleton(tmp_path: Path) -> Path:
    """Same shape as `synthetic_template_path` but with an extra eVars
    skeleton row for `campaign` at row 10 — used by the eVars-filtering
    test that needs `campaign` to be present as a pre-seeded id."""
    return _build_template(
        tmp_path / "synthetic_template_campaign.xlsx",
        evars_skeleton=[
            ["1", "evar1", "Pre-seeded eVar 1", "Pre-seeded desc 1"],
            ["2", "evar2", "Pre-seeded eVar 2", "Pre-seeded desc 2"],
            ["3", "evar37", "Skeleton eVar 37", "Skeleton desc 37"],
            ["4", "campaign", "Pre-seeded campaign", "Pre-seeded campaign desc"],
        ],
    )
