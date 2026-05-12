"""Pure header/column locators for the template-fill Excel writer.

No I/O, no openpyxl writes. Given a `Workbook` and an expected anchor
metadata record, return either a `ResolvedSheet` (header row found + column
map built) or `None` (caller logs + skips). See spec §3.4.

Case-insensitive `B4` section-title check tolerates Adobe template authors
renaming the section title. Header scan is rows 5–10 only; deeper headers
are intentionally unsupported (real templates put the header at row 6).
v1.16.0."""

from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook


@dataclass(frozen=True, slots=True)
class SheetAnchor:
    """Expected anchor record for one of the template's data sheets."""

    sheet_name: str
    section_title: str
    id_marker: str = "ID"


@dataclass(frozen=True, slots=True)
class ResolvedSheet:
    """Result of a successful anchor resolution. 1-indexed (openpyxl convention)."""

    sheet_name: str
    header_row: int
    columns: dict[str, int]
    first_data_row: int


# Per spec §3.5. Glossary handled separately (single-cell C2 write, no anchors).
ANCHORS: dict[str, SheetAnchor] = {
    "evars": SheetAnchor("eVars", "eVars"),
    "props": SheetAnchor("props", "Props"),
    "events": SheetAnchor("custom events (metrics)", "Custom Events (Metrics)"),
    "metrics_segments": SheetAnchor("metrics-segments", "Metrics - Segments"),
}


def resolve_sheet(wb: Workbook, anchor: SheetAnchor) -> ResolvedSheet | None:
    """Return a ResolvedSheet on success; None when any check fails."""
    if anchor.sheet_name not in wb.sheetnames:
        return None
    ws = wb[anchor.sheet_name]
    b4_value = str(ws["B4"].value or "").strip().lower()
    if b4_value != anchor.section_title.strip().lower():
        return None
    for row in range(5, 11):
        if str(ws.cell(row=row, column=2).value or "").strip() == anchor.id_marker:
            columns: dict[str, int] = {}
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    continue
                columns[str(cell_value).strip()] = col
            return ResolvedSheet(
                sheet_name=anchor.sheet_name,
                header_row=row,
                columns=columns,
                first_data_row=row + 1,
            )
    return None
