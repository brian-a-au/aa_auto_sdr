"""Template-fill Excel writer (v1.16.0).

Opens an existing Adobe BRD/SDR `.xlsx` template, locates each data sheet by
name + section anchor, finds the header row by content, and fills component
data into the rows below. Preserves styles, cross-sheet formulas, defined
names, column widths, page setup, and every cell not explicitly written.

Heavy import (`openpyxl`) is deferred to method scope so the registry can be
bootstrapped on fast paths without paying the import cost. Mirrors the
deferred-pandas pattern in `excel.py`.

Self-registers under the `excel-template` format key on import.

Read-only AA 2.0 invariant: `openpyxl` writes target local `.xlsx` files only,
never the Adobe Analytics API."""

from __future__ import annotations

import logging
import time
from collections.abc import Callable, Iterable
from pathlib import Path
from typing import Any

from aa_auto_sdr.output._template_anchors import ANCHORS, SheetAnchor, resolve_sheet
from aa_auto_sdr.output.registry import register_writer
from aa_auto_sdr.sdr.document import SdrDocument

logger = logging.getLogger(__name__)


class ExcelTemplateWriter:
    extension = ".xlsx"

    def __init__(self) -> None:
        self.template_path: Path | None = None
        self.organization: str | None = None

    # ---- shared sheet-fill engine -------------------------------------------

    def _fill_by_id_sheet(
        self,
        wb,
        *,
        anchor: SheetAnchor,
        id_column_header: str,
        components: Iterable[Any],
        get_id: Callable[[Any], str],
        column_map: dict[str, Callable[[Any], Any]],
    ) -> None:
        """Locate a sheet, build the column index, fill rows by id-match.

        - `id_column_header`: the column whose value is the row's component id
          (e.g. 'Analytics Variable' for dimensions, 'Event' for events).
        - `column_map`: header text → getter that returns the value to write
          for a given component. Headers absent from the sheet's column map
          are silently skipped (the template may not expose every column).
        - Non-empty rule (§3.7): only write when the value is non-empty after
          stripping; never blank-overwrite an existing styled cell.
        - Soft cap: never append past max_row + 50 rows.
        """
        resolved = resolve_sheet(wb, anchor)
        if resolved is None:
            logger.warning(
                "template_sheet_skipped sheet=%s reason=missing_or_unanchored",
                anchor.sheet_name,
                extra={
                    "sheet": anchor.sheet_name,
                    "reason": "missing_or_unanchored",
                },
            )
            return

        ws = wb[resolved.sheet_name]
        if id_column_header not in resolved.columns:
            logger.warning(
                "template_sheet_skipped sheet=%s reason=no_id_column",
                anchor.sheet_name,
                extra={"sheet": anchor.sheet_name, "reason": "no_id_column"},
            )
            return

        id_col = resolved.columns[id_column_header]
        wanted = {get_id(c).lower(): c for c in components}
        if not wanted:
            return

        rows_matched = 0
        for row in range(resolved.first_data_row, ws.max_row + 1):
            existing_id = ws.cell(row=row, column=id_col).value
            if existing_id is None:
                continue
            key = str(existing_id).strip().lower()
            if key not in wanted:
                continue
            comp = wanted.pop(key)
            _write_row(ws, row=row, columns=resolved.columns, column_map=column_map, comp=comp)
            rows_matched += 1

        # Append path for components the template didn't pre-seed.
        # Soft cap bounds the work at +50 rows past max_row; anything beyond
        # is dropped with a warning (Task 8 emits the warning separately).
        rows_appended = 0
        soft_cap = ws.max_row + 50
        append_row = ws.max_row + 1
        remaining = list(wanted.values())
        for comp in remaining:
            if append_row > soft_cap:
                break
            ws.cell(row=append_row, column=id_col, value=get_id(comp))
            _write_row(
                ws,
                row=append_row,
                columns=resolved.columns,
                column_map=column_map,
                comp=comp,
            )
            append_row += 1
            rows_appended += 1
        rows_dropped = len(remaining) - rows_appended

        logger.info(
            "template_sheet_filled sheet=%s rows_matched=%d rows_appended=%d",
            anchor.sheet_name,
            rows_matched,
            rows_appended,
            extra={
                "sheet": anchor.sheet_name,
                "rows_matched": rows_matched,
                "rows_appended": rows_appended,
            },
        )
        if rows_dropped > 0:
            logger.warning(
                "template_sheet_clipped sheet=%s rows_dropped=%d soft_cap=%d",
                anchor.sheet_name,
                rows_dropped,
                soft_cap,
                extra={
                    "sheet": anchor.sheet_name,
                    "rows_dropped": rows_dropped,
                    "soft_cap": soft_cap,
                },
            )

    # ---- dimensions ---------------------------------------------------------

    def _fill_dimensions(self, wb, doc: SdrDocument) -> None:
        evar_components = [d for d in doc.dimensions if d.id.lower().startswith("evar") or d.id == "campaign"]
        prop_components = [
            d for d in doc.dimensions if d.id.lower().startswith("prop") or d.id in {"pageName", "linkName"}
        ]
        column_map: dict[str, Callable[[Any], Any]] = {
            "Analytics Variable": lambda d: d.id,
            "Variable Name": lambda d: d.name,
            "Variable Description": lambda d: d.description,
        }
        self._fill_by_id_sheet(
            wb,
            anchor=ANCHORS["evars"],
            id_column_header="Analytics Variable",
            components=evar_components,
            get_id=lambda d: d.id,
            column_map=column_map,
        )
        self._fill_by_id_sheet(
            wb,
            anchor=ANCHORS["props"],
            id_column_header="Analytics Variable",
            components=prop_components,
            get_id=lambda d: d.id,
            column_map=column_map,
        )

    def _fill_metrics(self, wb, doc: SdrDocument) -> None:
        custom_events = [m for m in doc.metrics if m.id.lower().startswith("event")]
        column_map: dict[str, Callable[[Any], Any]] = {
            "Event": lambda m: m.id,
            "Event Name": lambda m: m.name,
            "Event Description": lambda m: m.description,
        }
        self._fill_by_id_sheet(
            wb,
            anchor=ANCHORS["events"],
            id_column_header="Event",
            components=custom_events,
            get_id=lambda m: m.id,
            column_map=column_map,
        )

    def _fill_glossary_org(self, wb, doc: SdrDocument) -> None:
        """Write the org name to Glossary!C2 — the source-of-truth cell
        every other sheet's C2 formula references. Skips if 'Glossary' is
        absent (defensive — real Adobe templates always include it)."""
        if "Glossary" not in wb.sheetnames:
            return
        org = self.organization or doc.report_suite.name
        wb["Glossary"]["C2"] = org

    def write(self, doc: SdrDocument, output_path: Path) -> list[Path]:
        if self.template_path is None:
            raise RuntimeError(
                "ExcelTemplateWriter dispatched without template_path; CLI validator should have caught this.",
            )
        from openpyxl import load_workbook  # method-scoped lazy import

        started = time.monotonic()
        target = output_path if output_path.suffix == self.extension else output_path.with_suffix(self.extension)
        wb = load_workbook(self.template_path)  # data_only=False; preserve formulas

        logger.info(
            "template_load path=%s sheets=%d",
            str(self.template_path),
            len(wb.sheetnames),
            extra={"path": str(self.template_path), "sheets": len(wb.sheetnames)},
        )

        self._fill_glossary_org(wb, doc)
        self._fill_dimensions(wb, doc)
        self._fill_metrics(wb, doc)

        # Subsequent tasks add: _fill_metrics_segments.

        target.parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)

        duration_ms = int((time.monotonic() - started) * 1000)
        logger.info(
            "output_write format=excel-template output_path=%s count=1 duration_ms=%s",
            str(target),
            duration_ms,
            extra={
                "format": "excel-template",
                "output_path": str(target),
                "count": 1,
                "duration_ms": duration_ms,
                "rsid": doc.report_suite.rsid,
            },
        )
        return [target]


def _write_row(
    ws,
    *,
    row: int,
    columns: dict[str, int],
    column_map: dict[str, Callable[[Any], Any]],
    comp: Any,
) -> None:
    """Write one component's values into a row, honoring the non-empty rule.

    Never writes None / empty over an existing styled cell — preserves the
    template's example content when the API has no data for that field.
    Headers absent from `columns` (i.e. not in this sheet) are silently
    skipped."""
    for header, getter in column_map.items():
        col = columns.get(header)
        if col is None:
            continue
        value = getter(comp)
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        ws.cell(row=row, column=col, value=value)


register_writer("excel-template", ExcelTemplateWriter())
